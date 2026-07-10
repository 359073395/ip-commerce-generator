import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "outputs" / "account_video_deep_analysis" / "video-deep-analysis.json"
TARGET = ROOT / "outputs" / "account_video_deep_analysis" / "video-deep-analysis.xlsx"


def join_values(value, limit=None):
    if value is None:
        return ""
    if isinstance(value, list):
        items = value[:limit] if limit else value
        return "\n".join(str(item) for item in items)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def style_sheet(ws, widths):
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_rows(ws, headers, rows, mapper):
    ws.append(headers)
    for row in rows:
        ws.append(mapper(row))


def main():
    data = json.loads(SOURCE.read_text(encoding="utf-8-sig"))
    TARGET.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Videos"
    video_headers = [
        "序号",
        "账号ID",
        "账号",
        "行业组",
        "标题",
        "链接",
        "数据",
        "发布时间",
        "时长秒",
        "钩子机制",
        "文案结构",
        "画面语法",
        "转化信号",
        "可复用公式",
        "证据等级",
        "章节摘要",
        "评论片段",
    ]
    append_rows(
        ws,
        video_headers,
        data["rows"],
        lambda row: [
            row.get("index"),
            row.get("accountId"),
            row.get("accountName"),
            row.get("group"),
            row.get("title"),
            row.get("url"),
            row.get("metricText"),
            row.get("publishDate"),
            row.get("durationSec"),
            join_values(row.get("hookMechanics")),
            row.get("structureFormula"),
            row.get("visualGrammar"),
            join_values(row.get("conversionSignals")),
            row.get("reusableFormula"),
            row.get("evidenceLevel"),
            join_values(row.get("chapterEvidence"), limit=12),
            join_values(row.get("commentEvidence"), limit=10),
        ],
    )
    style_sheet(ws, [8, 10, 22, 18, 46, 45, 12, 18, 10, 22, 46, 42, 24, 42, 24, 60, 42])

    summary_ws = wb.create_sheet("GroupSummary")
    summary_headers = ["行业组", "账号", "视频数", "章节摘要数", "高频钩子", "核心公式", "代表视频"]
    append_rows(
        summary_ws,
        summary_headers,
        data["summaries"],
        lambda row: [
            row.get("group"),
            join_values(row.get("accounts")),
            row.get("videos"),
            row.get("chapterCount"),
            join_values(row.get("topHooks")),
            join_values(row.get("coreFormulas")),
            join_values(row.get("representativeVideos")),
        ],
    )
    style_sheet(summary_ws, [18, 30, 10, 12, 30, 60, 60])

    method_ws = wb.create_sheet("MethodCards")
    method_headers = ["方法卡ID", "标题", "适用模块", "方法", "场景", "必填信息", "输出骨架", "示例", "关键词"]
    append_rows(
        method_ws,
        method_headers,
        data["methodBlocks"],
        lambda row: [
            row.get("id"),
            row.get("title"),
            join_values(row.get("moduleIds")),
            join_values(row.get("methods")),
            join_values(row.get("scenarios")),
            join_values(row.get("requiredInputs")),
            join_values(row.get("outputTemplate")),
            row.get("example"),
            join_values(row.get("keywords")),
        ],
    )
    style_sheet(method_ws, [38, 26, 20, 28, 28, 30, 34, 58, 32])

    wb.save(TARGET)
    print(TARGET)


if __name__ == "__main__":
    main()
